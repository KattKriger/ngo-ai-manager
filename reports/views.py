from .models import Report
from django.shortcuts import render, redirect
from .forms import ReportForm
from django.db.models import Sum
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl import load_workbook
from django.shortcuts import get_object_or_404


def report_list(request):

    chart_type = request.GET.get('chart', 'attendance')

    if chart_type == 'deaths':
        chart_data = (
            Report.objects
            .values('year')
            .annotate(total=Sum('deaths'))
            .order_by('year')
        )
    else:
        chart_data = (
            Report.objects
            .values('year')
            .annotate(total=Sum('internal_attendance') + Sum('external_attendance'))
            .order_by('year')
        )

    selected_year = request.GET.get('year')

    reports = Report.objects.all()

    attendance_by_year = (
        Report.objects
        .values('year')
        .annotate(total=Sum('internal_attendance') + Sum('external_attendance'))
        .order_by('year')
    )

    data = list(attendance_by_year)

    highest_year = None
    lowest_year = None
    trend = "Not enough data"

    if len(data) > 0:
        highest = max(data, key=lambda x: x['total'])
        lowest = min(data, key=lambda x: x['total'])

        highest_year = highest['year']
        lowest_year = lowest['year']

        if len(data) > 1:

            if data[-1]['total'] > data[0]['total']:
                trend = "Increasing"
            elif data[-1]['total'] < data[0]['total']:
                trend = "Decreasing"
            else:
                trend = "Stable"

    if chart_type == 'deaths':
        chart_title = "Deaths by Year"
    else:
        chart_title = "Attendance by Year"

    chart_data = (
        Report.objects
        .values('year')
        .annotate(total_attendance=Sum('internal_attendance') + Sum('external_attendance'))
        .order_by('year')
    )
    total_reports = reports.count()
    total_internal_attendance = reports.aggregate(
        Sum('internal_attendance')
        )['internal_attendance__sum']
    total_external_attendance = reports.aggregate(
        Sum('external_attendance')
        )['external_attendance__sum']
    total_deaths = reports.aggregate(
        Sum('deaths')
        )['deaths__sum']

    if selected_year:
            reports = reports.filter(year=selected_year)
        
    years = Report.objects.values_list(
        'year',
        flat=True
    ).distinct().order_by('year')

    recommendations = []
    if trend == "Increasing":
        recommendations.append("Attendance is gorwing. Consider increasing shelter resources.")
    if total_deaths and total_reports:
        average_deaths = total_deaths / total_reports

        if average_deaths > 10:
            recommendations.append("Death rate is high. Investigate health and veterinary processes.")
    if total_external_attendance and total_internal_attendance:
        recommendations.append("External attendance exceeds internal attendance. Review outreach programs.")

    attendance_values = [item['total_attendance'] for item in chart_data]
    predicted_next_year = None
    if len(attendance_values) >= 2:
        growth = (attendance_values[-1] - attendance_values[0]) / (len(attendance_values) - 1)
        predicted_next_year = int(attendance_values[-1] + growth)


    #TEMPLATE
    return render(request, 'reports/report_list.html', {
        'reports': reports,
        'years': years,
        'selected_year': selected_year,
        'total_reports': total_reports,
        'total_internal_attendance': total_internal_attendance,
        'total_external_attendance': total_external_attendance,
        'total_deaths': total_deaths,
        'highest_year': highest_year,
        'lowest_year': lowest_year,
        'trend': trend,
        'chart_data': chart_data,
        'recommendations': recommendations,
        'predicted_next_year': predicted_next_year,
        'chart_type': chart_type,
        'chart_title': chart_title
    })

def create_report(request):
    if request.method == 'POST':
        form = ReportForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('/')
    else:
        form = ReportForm()
    return render(
        request,
        'reports/create_report.html',
        {'form': form}
    )

def export_excel(request):

    reports = Report.objects.all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "NGO Reports"
    
    ws.append([
        'Year',
        'Month',
        'Internal Attendance',
        'External Attendance',
        'Deaths'
    ])
    
    for report in reports:
        ws.append([
            report.year,
            report.month,
            report.internal_attendance,
            report.external_attendance,
            report.deaths
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    response['Content-Disposition'] = ('attachment; filename=ngo_reports.xlsx')

    wb.save(response)
    
    return response

def import_excel(request):

    if request.method == 'POST':

        excel_file = request.FILES['excel_file']

        workbook = load_workbook(excel_file)

        sheet = workbook.active

        imported = 0
        skipped = 0
        
        for row in sheet.iter_rows(min_row=2, values_only=True):

            year = row[0]
            month = row[1]
            internal_attendance = row[2]
            external_attendance = row[3]
            deaths = row[4]

            existing = Report.objects.filter(
                year=year,
                month=month
            ).exists()

            if existing:
                skipped += 1
                continue

            imported += 1

            Report.objects.create(
                year=year,
                month=month,
                internal_attendance=internal_attendance,
                external_attendance=external_attendance,
                deaths=deaths
            )

        print(f"Imported: {imported}")
        print(f"Skipped: {skipped}")

        return redirect('/')
    
    return render(request, 'reports/import_excel.html')

def edit_report(request, report_id):
    
    report = get_object_or_404(
        Report,
        id=report_id
    )

    if request.method == 'POST':
    
        form = ReportForm(
            request.POST,
            instance=report
        )

        if form.is_valid():
            form.save()
            return redirect('/')
    else:
        form = ReportForm(
            instance=report
        )

    return render(
        request,
        'reports/create_report.html',
        {'form': form}
    )

def delete_report(request, report_id):
    
    report = get_object_or_404(
        Report,
        id=report_id
    )

    report.delete()

    return redirect('/')

def ai_assistant(request):

    answer = ""

    Reports = Report.objects.all()

    attendance_by_year = {}

    for report in Reports:

        total = (report.internal_attendance or 0) + (report.external_attendance or 0)
        attendance_by_year[report.year] = (attendance_by_year.get(report.year, 0) + total)

        highest_year = max(attendance_by_year, key=attendance_by_year.get)
        lowest_year = min(attendance_by_year, key=attendance_by_year.get)
        total_attendance = sum(attendance_by_year.values())
        total_deaths = sum((report.deaths or 0) for report in Reports)

    years = sorted(attendance_by_year.keys())
    predicted_next_year = 0

    if len(years) >= 2:
       attendance_values = list(attendance_by_year.values())
       predicted_next_year = None

    if len(attendance_values) >= 2:
        growth = (attendance_values[-1] - attendance_values[0]) / (len(attendance_values) - 1)
        predicted_next_year = int(attendance_values[-1] + growth)

    elif len(years) == 1:
        predicted_next_year = attendance_by_year[years[0]]



    if request.method == 'POST':
        question = request.POST.get('question',"").lower()

        
        if "deaths chart" in question:
             return redirect('/?chart=deaths')

        elif "attendance chart" in question:
             return redirect('/?chart=attendance')

        elif "best year" in question:
            answer = f"The best attendance year was {highest_year}"
        
        elif "lowest year" in question:
            answer = f"The lowest attendance year was {lowest_year}"

        elif "total attendance" in question:
            answer = f"The total attendance is {total_attendance}"

        elif "deaths" in question:
            answer = f"The total deaths recorded: {total_deaths}"

        elif "reports" in question:
            answer = f"There are {Reports.count()} reports"

        elif "forecast" in question:
            answer = f"Predicted attendance for next year: {predicted_next_year}"

        else:
            answer = "I don't understand the question yet."
            
    return render(request, 'reports/ai_assistant.html', {'answer': answer})
    